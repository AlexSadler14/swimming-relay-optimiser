"""
Swimming Relay Optimiser -- CLI entry point.

Usage:
    python main.py swimmers.csv --club "Trafford Metro SC" --output results.xlsx
    Options:
      --allow-72plus   include the 72+ (Senior Age Group) category (off by default)
      --max-relays N   cap every swimmer at N relays (overridden by a per-swimmer column)

Expected CSV columns:
    name, age, gender,
    max_relays,                      (optional: max relays this swimmer may swim; blank = no limit)
    mens_4x50_free, mens_4x50_free_time,
    womens_4x50_free, womens_4x50_free_time,
    mixed_4x50_free, mixed_4x50_free_time,
    ... (same pattern for 4x100_free, 4x200_free)
    mens_4x50_medley, mens_4x50_medley_stroke, mens_4x50_medley_time,
    womens_4x50_medley, womens_4x50_medley_stroke, womens_4x50_medley_time,
    mixed_4x50_medley, mixed_4x50_medley_stroke, mixed_4x50_medley_time,
    ... (same pattern for 4x100_medley)

Y/N columns: y or n (case-insensitive). Blank treated as n.
Time format: seconds (28, 28.5, 28.50) or M:SS (1:05, 1:05.23). Decimals optional.
Medley stroke: back, breast, fly, or free. Multiple: free,back
Medley time: matches stroke order. Multiple: 22, 26.5
"""
import re
import sys
import csv
import os
import argparse

# Allow imports from src/
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "src"))

from models import Swimmer
from record_fetcher import RecordsFetcher
import optimiser
import reporter
import config
from config import EVENT_GENDER_COLS, RELAY_EVENTS

# Accepted time formats (decimals optional, 1 or 2 d.p.):
#   "28"  "28.5"  "28.50"  "1:05"  "1:05.2"  "1:05.23"  "10:05.23"
_TIME_RE = re.compile(r"^(\d+:)?\d{1,2}(\.\d{1,2})?$")

# Values that mean "not entering this event" -- no warning needed
_BLANK_VALUES = {"", "-", "n/a", "none", "null", "no", "n"}

VALID_STROKES = {"back", "breast", "fly", "free"}


def parse_time(value: str, swimmer_name: str, field: str) -> "tuple[float | None, str | None]":
    """
    Parse a time string to float seconds.
    Accepts only M:SS.ss or SS.ss (decimal seconds with exactly 2 d.p.).
    Any other non-blank value is flagged and treated as opted-out.
    """
    if value is None:
        return None, None
    v = str(value).strip().lower()
    if v in _BLANK_VALUES:
        return None, None
    if not _TIME_RE.match(v):
        return None, (f"  [INPUT WARNING] {swimmer_name} -- {field}: "
                      f"unrecognised format {value!r}  "
                      f"(expected seconds or M:SS, e.g. 28, 28.5, 28.50 or 1:05.23) -- treated as blank")
    parts = v.split(":")
    try:
        secs = int(parts[0]) * 60 + float(parts[1]) if len(parts) == 2 else float(parts[0])
    except ValueError:
        return None, (f"  [INPUT WARNING] {swimmer_name} -- {field}: "
                      f"could not parse {value!r} -- treated as blank")

    # Sanity-check: relay split times should be between 10s and 20 minutes
    if not (10.0 <= secs <= 1200.0):
        return None, (f"  [INPUT WARNING] {swimmer_name} -- {field}: "
                      f"time {value!r} = {secs:.1f}s is outside the expected range "
                      f"(10s - 20min) -- treated as blank")
    return secs, None


def _secs_to_time_str(total_secs: float) -> str:
    """Convert a total-seconds float to M:SS.ss or SS.ss string."""
    mins = int(total_secs // 60)
    secs = total_secs - mins * 60
    return f"{mins}:{secs:05.2f}" if mins > 0 else f"{secs:.2f}"


def _excel_val_to_str(value) -> str:
    """
    Convert an openpyxl cell value to a plain string.

    Handles the common case where Excel auto-converts a swimming time like
    '1:05.23' or '00:31.00' into an internal time/datetime value.
    """
    import datetime
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (datetime.time, datetime.datetime)):
        t = value if isinstance(value, datetime.time) else value.time()
        total_secs = t.hour * 3600 + t.minute * 60 + t.second + t.microsecond / 1_000_000
        return _secs_to_time_str(total_secs)
    if isinstance(value, float):
        if 0 < value < 1:   # Excel time fraction (fraction of a day)
            return _secs_to_time_str(value * 86400)
        return f"{value:.2f}" if value != int(value) else str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _load_rows(path: str) -> "tuple[list[str], list[dict]]":
    """
    Return (fieldnames, rows) from a CSV or Excel (.xlsx) file.
    All values are plain strings ready for the existing parsing logic.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return [], []
        fieldnames = [str(h).strip().lower() if h is not None else ""
                      for h in all_rows[0]]
        rows = [
            {fieldnames[j]: _excel_val_to_str(v)
             for j, v in enumerate(row_vals) if j < len(fieldnames)}
            for row_vals in all_rows[1:]
        ]
        return fieldnames, rows
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            reader.fieldnames = [h.strip().lower() for h in (reader.fieldnames or [])]
            return reader.fieldnames, list(reader)


def load_swimmers(path: str) -> list:
    """Load swimmers from a CSV or Excel file, printing all input warnings before returning."""
    swimmers = []
    warnings = []

    _, rows = _load_rows(path)

    for i, row in enumerate(rows, 1):
        name = row.get("name", "").strip()
        if not name:
            continue

        row_warnings = []

        # --- age ---
        age_raw = row.get("age", "").strip()
        try:
            age = int(age_raw)
            if not (15 <= age <= 110):
                row_warnings.append(
                    f"  [INPUT WARNING] {name!r} -- age {age_raw!r} is outside "
                    f"the expected range (15-110)")
        except ValueError:
            warnings.append(f"  [INPUT ERROR]   row {i} ({name!r}): "
                             f"age {age_raw!r} is not a whole number -- row skipped")
            continue

        # --- gender ---
        gender = row.get("gender", "").strip().upper()
        if gender not in ("M", "F"):
            warnings.append(f"  [INPUT ERROR]   row {i} ({name!r}): "
                             f"gender must be M or F (got {row.get('gender','')!r}) -- row skipped")
            continue

        # --- max relays (optional per-swimmer cap; blank = no limit) ---
        max_relays = None
        max_relays_raw = row.get("max_relays", "").strip()
        if max_relays_raw:
            try:
                max_relays = int(max_relays_raw)
                if max_relays < 1:
                    row_warnings.append(
                        f"  [INPUT WARNING] {name} -- max_relays {max_relays_raw!r} "
                        f"must be 1 or more -- ignored (no limit)")
                    max_relays = None
            except ValueError:
                row_warnings.append(
                    f"  [INPUT WARNING] {name} -- max_relays {max_relays_raw!r} "
                    f"is not a whole number -- ignored (no limit)")

        # --- relay entries ---
        entries = {}
        for (event, gender_key), col_info in EVENT_GENDER_COLS.items():
            yn_col = col_info["yn"]
            yn_raw = row.get(yn_col, "").strip().lower()

            if yn_raw != "y":
                continue  # not entering

            is_medley = "stroke" in col_info
            time_col = col_info["time"]

            if is_medley:
                # Parse stroke(s) and time(s) -- may be comma-separated
                stroke_col = col_info["stroke"]
                stroke_raw = row.get(stroke_col, "").strip().lower()
                time_raw = row.get(time_col, "").strip()

                if not stroke_raw or not time_raw:
                    row_warnings.append(
                        f"  [INPUT WARNING] {name} -- {yn_col}: marked y but "
                        f"stroke or time is missing -- entry skipped")
                    continue

                stroke_parts = [s.strip() for s in stroke_raw.split(",")]
                time_parts = [t.strip() for t in time_raw.split(",")]

                if len(stroke_parts) != len(time_parts):
                    row_warnings.append(
                        f"  [INPUT WARNING] {name} -- {yn_col}: "
                        f"{len(stroke_parts)} stroke(s) but {len(time_parts)} time(s) -- entry skipped")
                    continue

                entry_list = []
                valid = True
                for stroke_str, time_str in zip(stroke_parts, time_parts):
                    if stroke_str not in VALID_STROKES:
                        row_warnings.append(
                            f"  [INPUT WARNING] {name} -- {stroke_col}: "
                            f"{stroke_str!r} not valid (back/breast/fly/free) -- entry skipped")
                        valid = False
                        break
                    secs, warn = parse_time(time_str, name, time_col)
                    if warn:
                        row_warnings.append(warn)
                    if secs is None:
                        valid = False
                        break
                    entry_list.append({"stroke": stroke_str, "time": secs})

                if valid and entry_list:
                    entries[(event, gender_key)] = entry_list
            else:
                # Freestyle: just time
                time_raw = row.get(time_col, "").strip()
                secs, warn = parse_time(time_raw, name, time_col)
                if warn:
                    row_warnings.append(warn)
                if secs is not None:
                    entries[(event, gender_key)] = [{"stroke": "free", "time": secs}]

        warnings.extend(row_warnings)
        swimmers.append(Swimmer(name=name, age=age, gender=gender,
                                entries=entries, max_relays=max_relays))

    # Print all warnings together after loading
    if warnings:
        print("  --- Input validation ---")
        for w in warnings:
            print(w)
        print()

    return swimmers


def _print_medley_coverage(swimmers: list):
    """Print stroke coverage for each medley event-gender to help spot gaps."""
    medley_events = [e for e in RELAY_EVENTS if "medley" in e]
    from config import GENDER_KEYS
    for event in medley_events:
        for gender_key in GENDER_KEYS:
            counts = {"back": 0, "breast": 0, "fly": 0, "free": 0}
            for sw in swimmers:
                for entry in sw.get_entries(event, gender_key):
                    stroke = entry["stroke"]
                    if stroke in counts:
                        counts[stroke] += 1
            total = sum(counts.values())
            if total > 0:
                parts = ", ".join(f"{s}={c}" for s, c in counts.items())
                missing = [s for s, c in counts.items() if c == 0]
                warning = f"  ** missing: {', '.join(missing)}" if missing else ""
                label = f"{gender_key} {event}"
                print(f"    {label:<25s}  {parts}{warning}")


def main():
    parser = argparse.ArgumentParser(description="Swimming Relay Optimiser")
    parser.add_argument("swimmers", help="Path to swimmers CSV or Excel file")
    parser.add_argument("--output", default="results.xlsx",
                        help="Output filename, saved in the output/ folder "
                             "(default: results.xlsx)")
    parser.add_argument("--club", default="",
                        help="Club name shown in the report header")
    parser.add_argument("--csv", action="store_true",
                        help="Also export a raw CSV alongside the Excel file")
    parser.add_argument("--allow-72plus", action="store_true",
                        help="Enable the 72+ (Senior Age Group) relay category. "
                             "Off by default; many competitions don't run it.")
    parser.add_argument("--max-relays", type=int, default=None, metavar="N",
                        help="Default cap on relays per swimmer. A per-swimmer "
                             "'max_relays' column in the input overrides this.")
    args = parser.parse_args()

    config.ALLOW_72_PLUS_CATEGORY = args.allow_72plus

    print(f"Loading swimmers from {args.swimmers!r}...")
    swimmers = load_swimmers(args.swimmers)
    if not swimmers:
        print("No swimmers loaded. Check your input file.")
        sys.exit(1)

    # Apply the global --max-relays default to anyone without an explicit cap.
    if args.max_relays is not None:
        for s in swimmers:
            if s.max_relays is None:
                s.max_relays = args.max_relays

    print(f"  Loaded {len(swimmers)} swimmers.")
    if not config.ALLOW_72_PLUS_CATEGORY:
        print("  72+ (Senior Age Group) category: OFF  (use --allow-72plus to enable)")
    print()

    print("  Medley stroke coverage:")
    _print_medley_coverage(swimmers)
    print()

    print("Loading records database...")
    fetcher = RecordsFetcher()
    print(f"  Loaded {len(fetcher.all_records())} records.\n")

    print("Running optimiser...")
    committed = optimiser.run(swimmers, fetcher)
    stats = optimiser.summary_stats(committed)
    print(f"  {stats['total_teams']} relay teams found.")
    print(f"  World records potentially broken:    {stats['world']}")
    print(f"  European records potentially broken: {stats['european']}")
    print(f"  British records potentially broken:  {stats['british']}\n")

    reporter.print_summary(committed, fetcher)
    reporter.print_fastest_reference(swimmers, fetcher)

    # Always save into the output/ folder, using just the filename the user gave.
    OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
    filename = os.path.basename(args.output)
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    xl_path = os.path.join(OUTPUT_DIR, filename)
    try:
        reporter.export_excel(committed, fetcher, xl_path, club_name=args.club, swimmers=swimmers)
    except PermissionError:
        # Almost always means the file is open in Excel -- save under a new name
        # instead of crashing so the user doesn't lose the run.
        base, ext = os.path.splitext(xl_path)
        from datetime import datetime
        fallback = f"{base}_{datetime.now():%H%M%S}{ext}"
        print(f"  [WARNING] Could not write {xl_path!r} -- is it open in Excel?")
        print(f"            Saving to {fallback!r} instead.")
        reporter.export_excel(committed, fetcher, fallback, club_name=args.club, swimmers=swimmers)
        xl_path = fallback

    if args.csv:
        csv_path = xl_path.replace(".xlsx", ".csv")
        reporter.export_csv(committed, fetcher, csv_path)


if __name__ == "__main__":
    main()
