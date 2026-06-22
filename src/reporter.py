"""
Formats and outputs the optimisation results.

Outputs:
  - Console summary (print_summary)
  - Excel report (export_excel) -- mirrors the Trafford Metro report style
  - CSV export (export_csv)
"""
import csv
import os
from scorer import Scorer
from record_fetcher import RecordsFetcher

RECORD_LEVELS = ["world", "european", "british"]

_GENDER_LABEL = {"men": "Men", "women": "Women", "mixed": "Mixed"}
_GENDER_CODE  = {"men": "M", "women": "F", "mixed": "X"}
_EVENT_LABEL  = {
    "4x50_free":    "4x50m Freestyle",
    "4x100_free":   "4x100m Freestyle",
    "4x200_free":   "4x200m Freestyle",
    "4x50_medley":  "4x50m Medley",
    "4x100_medley": "4x100m Medley",
}
_DISTANCE = {
    "4x50_free": 200, "4x100_free": 400, "4x200_free": 800,
    "4x50_medley": 200, "4x100_medley": 400,
}


# ---------------------------------------------------------------------------
# Console summary
# ---------------------------------------------------------------------------

def print_summary(committed: dict, fetcher: RecordsFetcher):
    scorer = Scorer(fetcher)

    def _level_rank(broken):
        for level in RECORD_LEVELS:
            if level in broken:
                return RECORD_LEVELS.index(level)
        return 99

    entries = sorted(committed.items(),
                     key=lambda kv: (_level_rank(kv[1]["records_broken"]), kv[1]["score"]),
                     reverse=False)
    record_breakers = [(k, v) for k, v in entries if v["records_broken"]]
    other_teams     = [(k, v) for k, v in entries if not v["records_broken"]]

    print("=" * 70)
    print("  SWIMMING RELAY OPTIMISER -- RESULTS")
    print("=" * 70)
    print(f"\n  Record-breaking opportunities: {len(record_breakers)}")
    print(f"  Other valid relay entries:      {len(other_teams)}")
    print()

    # ---- record breaks summary table ----
    if record_breakers:
        print("-" * 70)
        print("  RECORDS BROKEN AT A GLANCE")
        print("-" * 70)
        for level in RECORD_LEVELS:
            level_rows = []
            for (event, gender, age_cat), v in record_breakers:
                # Only show under the highest level broken — skip if a higher level was broken
                if not v["records_broken"] or v["records_broken"][0] != level:
                    continue
                team = v["team"]
                rec  = fetcher.get_record(level, event, age_cat, gender)
                margin = rec.time - team.total_time
                label  = f"{_GENDER_LABEL[gender]} {_EVENT_LABEL.get(event, event)} [{age_cat}]"
                level_rows.append((margin, label, team.format_time(), _fmt_time(rec.time)))
            if not level_rows:
                continue
            level_rows.sort(key=lambda r: r[0], reverse=True)
            print(f"\n  {level.upper()} RECORD{'S' if len(level_rows) > 1 else ''}")
            for margin, label, team_time, rec_time in level_rows:
                print(f"    {label:<44s}  by {margin:5.2f}s"
                      f"  (team {team_time}  /  record {rec_time})")
        print()

    for section_label, section in [("RECORD-BREAKING RELAY TEAMS", record_breakers),
                                    ("OTHER VALID RELAY ENTRIES",   other_teams)]:
        if not section:
            continue
        print("-" * 70)
        print(f"  {section_label}")
        print("-" * 70)
        for (event, gender, age_cat), v in section:
            team   = v["team"]
            broken = v["records_broken"]
            print(f"\n  {_GENDER_LABEL[gender]} {_EVENT_LABEL.get(event, event)}  [{age_cat}]")
            print(f"    Time:   {team.format_time()}")
            print(f"    Breaks: {', '.join(b.upper() for b in broken) if broken else '--'}")
            for level in RECORD_LEVELS:
                rec = fetcher.get_record(level, event, age_cat, gender)
                if rec is None:
                    continue
                diff = team.total_time - rec.time   # negative = would break
                sign = "v" if diff < 0 else "^"
                print(f"    {level.capitalize():<10s}: {sign} {abs(diff):.2f}s  (record {_fmt_time(rec.time)})")
            print("    Swimmers:")
            for leg in team.legs:
                print(f"      * {leg.swimmer.name:<20s} {leg.stroke:<8s} {_fmt_split(leg.split_time)}")
    print("=" * 70)

    # ---- swimmer participation ----
    swimmer_events = {}   # name -> list of slot label strings
    swimmer_obj    = {}   # name -> Swimmer
    for (event, gender, age_cat), v in committed.items():
        label = f"{_GENDER_LABEL[gender]} {_EVENT_LABEL.get(event, event)} [{age_cat}]"
        for leg in v["team"].legs:
            sw = leg.swimmer
            swimmer_events.setdefault(sw.name, []).append(label)
            swimmer_obj[sw.name] = sw

    if swimmer_events:
        sorted_swimmers = sorted(swimmer_events.items(),
                                 key=lambda x: (-len(x[1]), x[0]))
        print()
        print("=" * 70)
        print("  SWIMMER PARTICIPATION")
        print("=" * 70)
        for name, event_list in sorted_swimmers:
            sw    = swimmer_obj[name]
            count = len(event_list)
            print(f"\n  {name:<25s}  {count} {'event' if count == 1 else 'events'}")
            for ev in sorted(event_list):
                print(f"    * {ev}")
        print()
        print("=" * 70)


# ---------------------------------------------------------------------------
# Excel export  (mirrors Trafford Metro report style)
# ---------------------------------------------------------------------------

def export_excel(committed: dict, fetcher: RecordsFetcher, path: str, club_name: str = ""):
    """
    Generate an Excel workbook with one table per relay team, matching the
    layout of the Trafford Metro 'Record Potential Teams' report.

    Columns:
      Swimmer 1 | Swimmer 2 | Swimmer 3 | Swimmer 4 |
      Age Group (Total) | Gender | Relay Event | Distance |
      Estimated Time |
      GB Record | Diff GB | EUR Record | Diff EUR | World Record | Diff World |
      Rating
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                 numbers)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Relay Teams"

    # ---- colour palette ----
    C_HEADER_BG   = "1F497D"   # dark blue
    C_HEADER_FG   = "FFFFFF"
    C_SUBHEAD_BG  = "DCE6F1"   # light blue
    C_BREAK_BG    = "E2EFDA"   # pale green  — breaks record
    C_CLOSE2_BG   = "FFF2CC"   # pale yellow — within 2s
    C_CLOSE4_BG   = "FCE4D6"   # pale orange — within 4s
    C_DIFF_NEG    = "375623"   # dark green text for negative diff (faster)
    C_DIFF_POS    = "843C0C"   # dark red text for positive diff (slower)
    C_SPLIT_BG    = "F2F2F2"   # light grey for split rows

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _font(bold=False, color="000000", size=10):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    def _border():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    def _center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _diff_fmt(diff: float) -> str:
        sign = "-" if diff < 0 else "+"
        return f"{sign}{abs(diff):.2f}"

    scorer = Scorer(fetcher)
    row = 1

    # ---- title row ----
    title = f"{club_name + '  --  ' if club_name else ''}Masters Relay Optimiser  (Long Course)"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=17)
    c = ws.cell(row, 1, title)
    c.font      = _font(bold=True, color=C_HEADER_FG, size=13)
    c.fill      = _fill(C_HEADER_BG)
    c.alignment = _center()
    ws.row_dimensions[row].height = 22
    row += 1

    # ---- legend row ----
    legend = ("*** = under British Record     ** = within 2s of British Record"
              "     * = within 4s of British Record     (negative Diff = would break record)")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=17)
    c = ws.cell(row, 1, legend)
    c.font      = _font(size=9, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 14
    row += 2

    # Sort: record-breakers first (best level), then by total_time
    def _sort_key(kv):
        v = kv[1]
        broken = v["records_broken"]
        level_rank = next((RECORD_LEVELS.index(l) for l in RECORD_LEVELS if l in broken), 99)
        return (level_rank, v["team"].total_time)

    sorted_entries = sorted(committed.items(), key=_sort_key)

    # ---- records broken summary section ----
    record_breakers_list = [(k, v) for k, v in sorted_entries if v["records_broken"]]
    if record_breakers_list:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=17)
        c = ws.cell(row, 1, "  RECORDS BROKEN AT A GLANCE")
        c.font      = _font(bold=True, color=C_HEADER_FG, size=11)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        for level in RECORD_LEVELS:
            level_rows = []
            for (event, gender, age_cat), v in record_breakers_list:
                # Only show under the highest level broken
                if not v["records_broken"] or v["records_broken"][0] != level:
                    continue
                team   = v["team"]
                rec    = fetcher.get_record(level, event, age_cat, gender)
                margin = rec.time - team.total_time
                label  = f"{_GENDER_LABEL[gender]} {_EVENT_LABEL.get(event, event)}  [{age_cat}]"
                level_rows.append((margin, label, team.format_time(), _fmt_time(rec.time)))

            if not level_rows:
                continue

            level_rows.sort(key=lambda r: r[0], reverse=True)

            # Level sub-header
            label_text = f"  {level.upper()} RECORD{'S' if len(level_rows) > 1 else ''}"
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=17)
            c = ws.cell(row, 1, label_text)
            c.font      = _font(bold=True, size=10)
            c.fill      = _fill(C_SUBHEAD_BG)
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 16
            row += 1

            for margin, label, team_time, rec_time in level_rows:
                # Event + gender + age group: cols 1-8
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
                c = ws.cell(row, 1, label)
                c.font      = _font(size=10)
                c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
                c.border    = _border()

                # Team time: cols 9-11
                ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=11)
                c = ws.cell(row, 9, f"Team: {team_time}")
                c.font      = _font(bold=True, size=10)
                c.alignment = _center()
                c.border    = _border()

                # Record time: cols 12-14
                ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=14)
                c = ws.cell(row, 12, f"Record: {rec_time}")
                c.font      = _font(size=10)
                c.alignment = _center()
                c.border    = _border()

                # Margin: cols 15-17
                ws.merge_cells(start_row=row, start_column=15, end_row=row, end_column=17)
                c = ws.cell(row, 15, f"faster by {margin:.2f}s")
                c.font      = _font(bold=True, color=C_DIFF_NEG, size=10)
                c.fill      = _fill(C_BREAK_BG)
                c.alignment = _center()
                c.border    = _border()

                ws.row_dimensions[row].height = 18
                row += 1

        row += 1  # blank row between summary and team tables

    for (event, gender, age_cat), v in sorted_entries:
        team  = v["team"]
        broken = v["records_broken"]

        gb_rec    = fetcher.get_record("british",  event, age_cat, gender)
        eur_rec   = fetcher.get_record("european", event, age_cat, gender)
        world_rec = fetcher.get_record("world",    event, age_cat, gender)

        diff_gb    = (team.total_time - gb_rec.time)    if gb_rec    else None
        diff_eur   = (team.total_time - eur_rec.time)   if eur_rec   else None
        diff_world = (team.total_time - world_rec.time) if world_rec else None

        # Star rating based on GB diff
        if diff_gb is not None and diff_gb < 0:
            rating = "***"
            row_bg = C_BREAK_BG
        elif diff_gb is not None and diff_gb <= 2:
            rating = "**"
            row_bg = C_CLOSE2_BG
        elif diff_gb is not None and diff_gb <= 4:
            rating = "*"
            row_bg = C_CLOSE4_BG
        else:
            rating = ""
            row_bg = None

        # ---- event header ----
        event_title = (f"{_GENDER_LABEL[gender]} {_EVENT_LABEL.get(event, event)}  "
                       f"[{age_cat}]  |  {_GENDER_CODE[gender]}  |  "
                       f"Dist: {_DISTANCE.get(event, '')}m")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=17)
        c = ws.cell(row, 1, event_title)
        c.font      = _font(bold=True, color=C_HEADER_FG, size=10)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 16
        row += 1

        # ---- column headers ----
        headers = [
            "Swimmer 1", "Swimmer 2", "Swimmer 3", "Swimmer 4",
            "Comb.\nAge", "Gen.", "Event\nType", "Dist\n(m)",
            "Est.\nTime",
            "GB Record", "Diff\nGB",
            "EUR Record", "Diff\nEUR",
            "World Record", "Diff\nWorld",
            "Rating",
        ]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row, col, h)
            c.font      = _font(bold=True, color=C_HEADER_FG, size=9)
            c.fill      = _fill("2E74B5")
            c.alignment = _center()
            c.border    = _border()
        ws.row_dimensions[row].height = 28
        row += 1

        # ---- swimmer names row ----
        fill = _fill(row_bg) if row_bg else PatternFill()
        legs = team.legs
        name_row = row

        for i, leg in enumerate(legs):
            label = f"{leg.swimmer.name}\n({leg.swimmer.age})"
            c = ws.cell(row, i + 1, label)
            c.font      = _font(bold=True, size=10)
            c.fill      = fill
            c.alignment = _center()
            c.border    = _border()

        combined_age = sum(l.swimmer.age for l in legs)
        values = [
            combined_age,
            _GENDER_CODE[gender],
            _EVENT_LABEL.get(event, event).split("m")[1].strip(),  # "Freestyle" / "Medley"
            _DISTANCE.get(event, ""),
            team.format_time(),
            _fmt_time(gb_rec.time)    if gb_rec    else "N/A",
            _diff_fmt(diff_gb)        if diff_gb is not None else "N/A",
            _fmt_time(eur_rec.time)   if eur_rec   else "N/A",
            _diff_fmt(diff_eur)       if diff_eur  is not None else "N/A",
            _fmt_time(world_rec.time) if world_rec else "N/A",
            _diff_fmt(diff_world)     if diff_world is not None else "N/A",
            rating,
        ]
        for i, val in enumerate(values):
            col = 5 + i
            c = ws.cell(row, col, val)
            c.font      = _font(bold=(i == 4), size=10)  # bold for estimated time
            c.fill      = fill
            c.alignment = _center()
            c.border    = _border()

            # Colour the diff cells
            if i in (6, 8, 10) and isinstance(val, str) and val not in ("N/A", ""):
                is_neg = val.startswith("-")
                c.font = _font(bold=True, color=(C_DIFF_NEG if is_neg else C_DIFF_POS), size=10)

            if i == 11 and rating:  # Rating cell
                c.font = _font(bold=True, color=(C_DIFF_NEG if rating == "***" else "000000"), size=11)

        ws.row_dimensions[row].height = 28
        row += 1

        # ---- split times + stroke row ----
        for i, leg in enumerate(legs):
            stroke_label = f"{leg.stroke}\n{_fmt_split(leg.split_time)}"
            c = ws.cell(row, i + 1, stroke_label)
            c.font      = _font(size=9, color="404040")
            c.fill      = _fill(C_SPLIT_BG)
            c.alignment = _center()
            c.border    = _border()

        # Merge remaining cols on split row for record reference detail
        detail = ""
        if diff_gb is not None:
            detail += f"GB: {_fmt_time(gb_rec.time)}  "
        if diff_eur is not None:
            detail += f"EUR: {_fmt_time(eur_rec.time)}  "
        if diff_world is not None:
            detail += f"WR: {_fmt_time(world_rec.time)}"
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=16)
        c = ws.cell(row, 5, detail)
        c.font      = _font(size=8, color="595959")
        c.fill      = _fill(C_SPLIT_BG)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border    = _border()

        ws.row_dimensions[row].height = 26
        row += 2   # blank row between entries

    # ---- swimmer participation section ----
    swimmer_events_xl = {}   # name -> list of slot label strings
    swimmer_obj_xl    = {}   # name -> Swimmer
    for (event, gender, age_cat), v in committed.items():
        label = f"{_GENDER_LABEL[gender]} {_EVENT_LABEL.get(event, event)} [{age_cat}]"
        for leg in v["team"].legs:
            sw = leg.swimmer
            swimmer_events_xl.setdefault(sw.name, []).append(label)
            swimmer_obj_xl[sw.name] = sw

    if swimmer_events_xl:
        sorted_swimmers_xl = sorted(swimmer_events_xl.items(),
                                    key=lambda x: (-len(x[1]), x[0]))

        # Section header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=17)
        c = ws.cell(row, 1, "  SWIMMER PARTICIPATION")
        c.font      = _font(bold=True, color=C_HEADER_FG, size=11)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        # Column headers
        for col_idx, label in enumerate(["Swimmer", "Events", "Competing In", "Notes"], 1):
            spans = [(1, 3), (4, 5), (6, 13), (14, 17)]
            s, e = spans[col_idx - 1]
            ws.merge_cells(start_row=row, start_column=s, end_row=row, end_column=e)
            c = ws.cell(row, s, label)
            c.font      = _font(bold=True, color=C_HEADER_FG, size=9)
            c.fill      = _fill("2E74B5")
            c.alignment = _center()
            c.border    = _border()
        ws.row_dimensions[row].height = 18
        row += 1

        for name, event_list in sorted_swimmers_xl:
            sw    = swimmer_obj_xl[name]
            count = len(event_list)
            events_text = "\n".join(sorted(event_list))
            row_height  = max(18, 14 * count)

            # Swimmer name
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            c = ws.cell(row, 1, name)
            c.font      = _font(bold=True, size=10)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border    = _border()

            # Event count
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
            c = ws.cell(row, 4, f"{count}")
            c.font      = _font(bold=True, size=10)
            c.alignment = _center()
            c.border    = _border()

            # Events list
            ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=13)
            c = ws.cell(row, 6, events_text)
            c.font      = _font(size=9)
            c.alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=True, indent=1)
            c.border    = _border()

            ws.row_dimensions[row].height = row_height
            row += 1

    # ---- column widths ----
    col_widths = [18, 18, 18, 18, 7, 5, 10, 6, 9, 11, 7, 11, 7, 12, 8, 7]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- freeze panes ----
    ws.freeze_panes = "A4"

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)
    print(f"Excel report saved to {path}")


# ---------------------------------------------------------------------------
# CSV export (raw data, useful for further analysis)
# ---------------------------------------------------------------------------

def export_csv(committed: dict, fetcher: RecordsFetcher, path: str):
    rows = []
    for (event, gender, age_cat), v in committed.items():
        team   = v["team"]
        broken = v["records_broken"]
        gb_rec    = fetcher.get_record("british",  event, age_cat, gender)
        eur_rec   = fetcher.get_record("european", event, age_cat, gender)
        world_rec = fetcher.get_record("world",    event, age_cat, gender)

        for pos, leg in enumerate(team.legs, 1):
            rows.append({
                "event":          event,
                "gender":         gender,
                "age_category":   age_cat,
                "relay_time":     team.format_time(),
                "records_broken": "|".join(broken),
                "diff_gb":    f"{team.total_time - gb_rec.time:+.2f}"    if gb_rec    else "",
                "diff_eur":   f"{team.total_time - eur_rec.time:+.2f}"   if eur_rec   else "",
                "diff_world": f"{team.total_time - world_rec.time:+.2f}" if world_rec else "",
                "leg":            pos,
                "swimmer":        leg.swimmer.name,
                "age":            leg.swimmer.age,
                "swimmer_gender": leg.swimmer.gender,
                "stroke":         leg.stroke,
                "split_time":     f"{leg.split_time:.2f}",
            })

    if not rows:
        print("No results to export.")
        return

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    print(f"CSV exported to {path}")


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _fmt_time(secs: float) -> str:
    mins = int(secs // 60)
    s = secs - mins * 60
    return f"{mins}:{s:05.2f}"


def _fmt_split(secs: float) -> str:
    mins = int(secs // 60)
    s = secs - mins * 60
    return f"{mins}:{s:05.2f}" if mins else f"0:{s:05.2f}"
